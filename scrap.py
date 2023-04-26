import pandas as pd
import os
import string
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path
from openpyxl.styles import Side, Alignment, Border, Font
from tkinter import *
from tkinter.ttk import *
import shutil
from typing import List

URLS_FOLDER = "./urls/"
LOGS = "logs"
OUTPUT="./Excel-katastre"
START_INDEX = 5
End_Index_P= 5
End_Index_O = 5
XLS_FILENAME = f"katastre_tabulka_{datetime.datetime.date(datetime.datetime.now())}.xlsx"
PARCELL_C_HEADER = {
                "alias_index": "Parcelné číslo",
                "alias_volume": "Výmera v m²",
                "alias_fie_type": "Druh pozemku",
                "alias_process": "Spôsob využívania pozemku",
                "alias_prot_type": "Druh chránenej nehnuteľnosti",
                "alias_share": "Spoločná nehnuteľnosť",
                "alias_place": "Umiestnenie pozemku",
                "alias_law": "Druh právneho vzťahu",
                "alias_add_info": "Iné údaje",
                "alias_note": "Zákonné veci",
                "alias_num": "Číslo LV"
}
PARCELL_E_HEADER = {
                "alias_index": "Parcelné číslo",
                "alias_volume": "Výmera v m²",
                "alias_fie_type": "Druh pozemku",
                "alias_base": "Pôvodné katastrálne územie",
                "alias_share": "Spoločná nehnuteľnosť",
                "alias_place": "Umiestnenie pozemku",
                "alias_add_info": "Iné údaje",
                "alias_note": "Zákonné veci",
                "alias_num": "Číslo LV"
}
OWNER_HEADER = {"alias_index": "Index",
                "alias_name": "Titul, priezvisko, meno, rodné meno / Miesto trvalého pobytu / Sídlo",
                #"alias_place": "Miesto trvalého pobytu / Sídlo",
                "alias_birth": "Dátum narodenia/ ICO",
                "alias_share": "Spoluvlastnícky podiel",
                "alias_percentage": "Spoluvlastnícky podiel v percentách",
                "alias_ownership": "Titul nadobudnutia",
                "alias_add_info": "Iné údaje",
                "alias_more": "Poznámky",
                "alias_plomba": "Plomba",
                "alias_parcell": "Číslo LV"
}
DEFAULT_HEIGHT = 50
DEFAULT_WIDTH = 40
OWNER_WIDTH =  {"alias_index": 10, #Default is 40
                "alias_share": 15,
                "alias_percentage": 15,
}





def create_logfile():
    file = open(f"{LOGS}/{str(datetime.datetime.now()).replace(':', '_')}.txt", "a")
    return file


def get_table_from_url(url):
    dfs = pd.read_html(url)
    log_file.write(f"scraping url - {url} complete\n")
    return dfs


def parse_owners_table(df, parcel_number):
    new = df.groupby([OWNER_HEADER["alias_index"]])[OWNER_HEADER["alias_name"]].apply(';;'.join).reset_index()
    log_file.write(f"table 1 - groupby {OWNER_HEADER['alias_index']} on {OWNER_HEADER['alias_name']} complete\n")

    #new[OWNER_HEADER["alias_place"]] = new[OWNER_HEADER["alias_name"]].apply(lambda x: x.split(", Dátum narodenia: ")[0] if ", Dátum narodenia: " in x else x.split("IČO: ")[0])
    #new[OWNER_HEADER["alias_place"]] = new[OWNER_HEADER["alias_name"]].apply(lambda x: x.split(".,")[1] if ".," in x else ",".join(x.split(",")[1:]))
    new[OWNER_HEADER["alias_birth"]] = new[OWNER_HEADER["alias_name"]].apply(lambda x: "---".join(x.split(", Dátum narodenia: ")[1:]) if ", Dátum narodenia: " in x else "---".join(x.split("IČO: ")[1:]) )#.str.split(", Dátum narodenia: ",expand=True)
    log_file.write(f"table 1 - split {OWNER_HEADER['alias_name']} to {OWNER_HEADER['alias_birth']}\n")

    new2 = df.drop_duplicates(subset=OWNER_HEADER["alias_index"], keep="first")
    new2 = new2.loc[:, new2.columns.isin([OWNER_HEADER["alias_index"], OWNER_HEADER["alias_share"]])]
    log_file.write(f"table 2 - keep only {OWNER_HEADER['alias_share']}\n")
    
    result = new.join(new2.set_index(OWNER_HEADER["alias_index"]), on=OWNER_HEADER["alias_index"]) # Join both tables via index
    log_file.write("table 1 and table 2 merged\n")

    result[OWNER_HEADER["alias_percentage"]] = result[OWNER_HEADER["alias_share"]].apply(lambda x: int(x.split("/")[0])/int(x.split("/")[1]))
    #result[[OWNER_HEADER["alias_name"],OWNER_HEADER["alias_birth"]]] = result[OWNER_HEADER["alias_name"]].str.split(", Dátum narodenia: ",expand=True)
    result[[OWNER_HEADER["alias_birth"],OWNER_HEADER["alias_ownership"]]] = result[OWNER_HEADER["alias_birth"]].str.split(";;Titul nadobudnutia;;",expand=True)
    result[[OWNER_HEADER["alias_ownership"],OWNER_HEADER["alias_add_info"]]] = result[OWNER_HEADER["alias_ownership"]].str.split(";;Iné údaje;;",expand=True)
    result[[OWNER_HEADER["alias_add_info"],OWNER_HEADER["alias_more"]]] = result[OWNER_HEADER["alias_add_info"]].str.split(";;Poznámky;;",expand=True)
    result[OWNER_HEADER["alias_plomba"]] = result[OWNER_HEADER["alias_birth"]].apply(lambda x: x.split(";;Plomba")[1] if OWNER_HEADER["alias_plomba"] in x else "")
    result[OWNER_HEADER["alias_birth"]] = result[OWNER_HEADER["alias_birth"]].apply(lambda x: x.split(";;Plomba")[0])

    result[OWNER_HEADER["alias_name"]] = result[OWNER_HEADER["alias_name"]].apply(lambda x: x.split(", Dátum narodenia: ")[0] if ", Dátum narodenia: " in x else x.split("IČO: ")[0])
    #result[OWNER_HEADER["alias_name"]] = result[OWNER_HEADER["alias_name"]].apply(lambda x: x.split(".,")[0] if ".," in x else x.split(",")[0])
    result[OWNER_HEADER["alias_parcell"]] = parcel_number
    log_file.write("data split into columns\n")

    result[OWNER_HEADER["alias_index"]] = result[OWNER_HEADER["alias_index"]].astype(int)
    #result[OWNER_HEADER["alias_parcell"]] = result[OWNER_HEADER["alias_parcell"]].astype(int)
    result = result.sort_values(by=[OWNER_HEADER["alias_index"]], ascending=True)
    log_file.write("result sorted, data ready!\n")
    return result


def load_table_to_excel(result, sheet, Start_Index):
    global End_Index_P
    global End_Index_O
        
    c_columns = len(result.columns)
    table_header = result.columns.values.tolist()

    for _, row in result.iterrows():
        Start_Index += 1
        for i in range(c_columns):
            # i+1 cause offset to start at B
            sheet[f"{string.ascii_uppercase[i+1]}{Start_Index}"] = row[table_header[i]]
    if sheet.title == "Pozemnky":
        End_Index_P = Start_Index+1
    else: 
        End_Index_O = Start_Index+1
    return 


def apply_style(sheet, start_row, max_row, header):
    # change height of cells
    for row in range(start_row ,max_row+1):
        sheet.row_dimensions[row].height = DEFAULT_HEIGHT
    log_file.write("style - changed heights\n")

    # set default aligment, border and fonts
    ali = Alignment(horizontal='center', vertical='top', text_rotation=0, wrap_text=True, shrink_to_fit=True, indent=0)
    thin = Border(right=Side(style='thin'),bottom=Side(style='thin'),top=Side(style='thin'),left=Side(style='thin'))
    font = Font(name='Calibri', size=10, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='000000')
    red_font = Font(name='Calibri', size=10, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF0000')
 
    for i, key in enumerate(header):
        col = string.ascii_uppercase[i+1]
        for row in range(start_row ,max_row+1):
            cell = sheet[f"{col}{row}"]
            cell.alignment = ali
            cell.border = thin
            cell.font = font
            
            if key == "alias_percentage": #TODO may cause bugs
                cell.number_format = '0.000%'
            if key == "alias_plomba" and row != start_row: #TODO may cause bugs
                cell.font = red_font

            # make thick header/ frame
            if i == 0 or i == len(header)-1:
                bot=top = Side(style='thin')
                if row == start_row:
                    top = Side(style='thick')
                    bot = Side(style='thick')
                elif row == max_row:
                    bot=Side(style='thick')
                if i == 0:
                    cell.border = Border(left=Side(style='thick'), top=top, bottom=bot, right=Side(style='thin'))
                else:
                    cell.border = Border(right=Side(style='thick'), top=top, bottom=bot, left=Side(style='thin'))
                    

            elif row == start_row:
                cell.border = Border(right=Side(style='thin'),bottom=Side(style='thick'),top=Side(style='thick'),left=Side(style='thin'))
            elif row == max_row:
                cell.border = Border(right=Side(style='thin'),bottom=Side(style='thick'),top=Side(style='thin'),left=Side(style='thin'))
    return


def change_width(sheet, col, key, wdict):
    sheet.column_dimensions[f"{col}"].width = wdict[key] if key in wdict else DEFAULT_WIDTH
    return


def create_header(sheet, header, wdict):
    if sheet.title == "Pozemnky":
        End_Index = End_Index_P
    else: 
        End_Index = End_Index_O

    for e, key in enumerate(header):
        # i+1 cause offset to start at B
        sheet[f"{string.ascii_uppercase[e+1]}{End_Index}"] = header[key]

        #change column width
        change_width(sheet, string.ascii_uppercase[e+1], key, wdict)
    return


def create_excel():
    if Path(XLS_FILENAME).exists():
        workbook = load_workbook(filename=XLS_FILENAME, data_only = False)
    else:
        workbook = Workbook()
    return workbook


def parse_owners(table, url_number, wb):
    table.columns =[OWNER_HEADER["alias_index"], OWNER_HEADER["alias_name"], OWNER_HEADER["alias_share"]]
    table = table.tail(-1)
    filt_table = parse_owners_table(table, url_number)
    log_file.write("filtered owner table complete\n")

    sheet = wb.active # TODO rewrite to name sheet
    create_header(sheet, OWNER_HEADER, OWNER_WIDTH)
    log_file.write("header created\n")

    start_table_index = End_Index_O
    load_table_to_excel(filt_table, sheet, End_Index_O)
    log_file.write("Owners parsed")

    apply_style(sheet, start_table_index, End_Index_O, OWNER_HEADER)
    log_file.write("styles applied\n")

    
def parse_plot_table(table: pd.DataFrame, header: str) -> pd.DataFrame:
    if header == PARCELL_E_HEADER:
        colls = list(header.values())[:6]
    elif header == PARCELL_C_HEADER:
        colls = list(header.values())[:8]
    table.columns = colls
    table = table.tail(-1)
    table = table.copy()
    table[header["alias_add_info"]] = ""

    add_info = ""
    for index, row in table.iterrows():
        if index == 1:
            last_row = row
            continue
        if row[header["alias_index"]] == row[header["alias_volume"]]: #idk some more info
            add_info += (";;" if len(add_info) != 0 else "") + str(row[header["alias_index"]]) 
        else:
            last_row[header["alias_add_info"]] = add_info
            add_info = ""
            last_row = row

    last_row[header["alias_add_info"]] = add_info
    return table


def refine_plot_columns(table: pd.DataFrame, header: str, url_number: str) -> pd.DataFrame:
    table[[header["alias_note"], header["alias_add_info"]]] = table[header["alias_add_info"]].str.split("Iné údaje: ",expand=True)
    table[header["alias_note"]] = table[header["alias_note"]].apply(lambda x: x.replace(";;", ""))
    table[header["alias_num"]] = url_number
    table = table[pd.to_numeric(table["Výmera v m²"], errors='coerce').notnull()]
    return table

def parse_plots(table: pd.DataFrame, url_number: str, wb: Workbook, header: str) -> None:
    filt_table = parse_plot_table(table, header)
    log_file.write("plots filtered\n")

    final_table = refine_plot_columns(filt_table, header, url_number)
    log_file.write("plots finalized\n")

    if "Pozemnky" in wb.sheetnames:
        sheet = wb["Pozemnky"]
    else:
        sheet = wb.create_sheet("ploty")
        sheet.title = "Pozemnky"

    create_header(sheet, header, {})
    log_file.write("header created\n")

    start_table_index = End_Index_P
    load_table_to_excel(final_table, sheet, End_Index_P)
    log_file.write("plots parsed")

    apply_style(sheet, start_table_index, End_Index_P, header)
    log_file.write("styles applied\n")
    log_file.write("raw plot table complete\n")


def parse_correct_tables(dfs: List[pd.DataFrame], url_number: str, wb: Workbook) -> None:
    for table in dfs:
        fst = table.astype(str).agg(';;'.join, axis=1).iloc[0] #TODO may casue errors
        if "Spoluvlastnícky podiel" in fst:
            parse_owners(table, url_number, wb)
        elif "Pôvodné katastrálne územie" in fst:
            parse_plots(table, url_number, wb, PARCELL_E_HEADER)
        elif "Spôsob využívania pozemku" in fst:
            parse_plots(table, url_number, wb, PARCELL_C_HEADER)

    
    log_file.write(f"file: {url} scraped to excel\n")


def program() -> None:  #TODO maybe split into more fnct later
    global log_file
    global START_INDEX
    global End_Index_O
    global End_Index_P
    global url

    log_file = create_logfile()
    if os.path.exists(OUTPUT):
        shutil.rmtree(OUTPUT)
    os.mkdir(OUTPUT)
    

    file_count = 0
    for _, _, files in os.walk(URLS_FOLDER):
        file_count += len(files)

    file_no=1
    bar['value'] = 0
    for dir_lv in os.scandir(URLS_FOLDER):
        dir_name = os.path.basename(dir_lv)
        print(dir_name)
        wb = create_excel()

        for url in os.scandir(dir_lv):
            bar['value'] += (1/file_count)*100
            percent.set(f"{round(((file_no/file_count)*100), 2)}%")
            text.set(f"{file_no}/{file_count} spravenych")
            window.update_idletasks()

            file_no += 1
            print(" |"+str(file_no) + "/" + str(file_count))
            df = get_table_from_url(url)
            parse_correct_tables(df, str(url).split("- ")[1].split(".html")[0], wb)
        wb.save(filename=f"{OUTPUT}/{dir_name}-{XLS_FILENAME}")
        START_INDEX = 5
        End_Index_P= 5
        End_Index_O = 5
    log_file.close()
    

def delete_results() -> None:
    shutil.rmtree(OUTPUT)


def UI_init() -> None:
    global text
    global percent
    global bar
    global window

    window = Tk()
    percent=StringVar()
    text=StringVar()

    bar = Progressbar(window, orient=HORIZONTAL, length=300)
    bar.pack(pady=10)

    Label(window, textvariable=percent).pack()
    Label(window, textvariable=text).pack()

    Button(window, text="spracovat tabulky", command=program).pack()
    Button(window, text="odstrániť tabulky", command=delete_results).pack()

    window.mainloop()


if __name__ == "__main__":
    UI_init()
