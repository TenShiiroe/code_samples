from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
import os
import shutil
import datetime
from pathlib import Path
from typing import Dict, Tuple, List

LOGS = "logs"
OUTPUT="./Excel-katastre"
START_INDEX = 5
COMP_FOLDER = f"{OUTPUT}/compare"
BASE_FILE = f"{COMP_FOLDER}/XXXXX"
COMP_FILE = f"{COMP_FOLDER}/YYYYYY"



OUTPUT = "./Porovnanie"
LOGS = LOGS
XLS_FILENAME = f"katastre_tabulka_{datetime.datetime.date(datetime.datetime.now())}.xlsx"

def program():
    """
    compares rows and creates excel with symetric difference of data
    """
    if os.path.exists(OUTPUT):
        shutil.rmtree(OUTPUT)
    os.mkdir(OUTPUT)
    shutil.copyfile(COMP_FILE, "comp_temp.xlsx")
    shutil.copyfile(BASE_FILE, "base_temp.xlsx")
    new_file = load_excel("comp_temp.xlsx")
    base_file = load_excel("base_temp.xlsx")
    base_all_data = get_all_rows(base_file)
    comp_all_data = get_all_rows(new_file)
    uniq_base = data_comparer(base_all_data, comp_all_data)
    uniq_comp = data_comparer(comp_all_data, base_all_data)
    wb = Workbook()
    last = write_data(uniq_base, Font(color="0d78d6"), wb, 5) # Blue - old not in new
    write_data(uniq_comp, Font(color="2aed18"), wb, last) # green - new not in old
    #comp_wb_lines(new_file, all_data)  # Useless
    wb.save(f"{OUTPUT}/result.xlsx")
    os.remove("comp_temp.xlsx")
    os.remove("base_temp.xlsx")


def write_data(dataset:Dict[str, List[Tuple[Cell,...]]], font: Font, wb: Workbook, start_pos: int) -> int:
    """
    writes data(<dataset>) into <wb> while using <font> on every row and starting from <start_pos>
    """
    sheet = wb.active
    assert sheet is not None
    sheet.title = "Vlastnici"


    for sheet_n in dataset.keys():
        print(sheet_n, wb.sheetnames, wb.worksheets)
        if sheet_n == "Sheet" and "Vlastnici" in wb.sheetnames:
            sheet = wb["Vlastnici"]
        elif sheet == "Sheet":
            sheet = wb.create_sheet("Vlastnici")
        elif sheet_n == "Pozemnky" and "Pozemnky" in wb.sheetnames:
            sheet = wb["Pozemnky"]
        else:
            sheet = sheet = wb.create_sheet("Pozemnky")
            
        for row in dataset[sheet_n]:
            for col in range(len(row)):
                sheet[f"{chr(col+65)}{start_pos}"] = row[col]  # +65 cause Ascii A
                sheet[f"{chr(col+65)}{start_pos}"].font = font
            start_pos += 1
    return start_pos


def data_comparer(data1: Dict[str, List[Tuple[Cell,...]]], data2: Dict[str, List[Tuple[Cell,...]]]) -> Dict[str, List[Tuple[Cell,...]]]:
    """
    compares rows and returns rows from data1 that are not in data2
    """
    temp = dict()
    for sheet in data1.keys():
        if sheet in data2.keys():
            temp[sheet] = [x for x in data1[sheet] if x not in data2[sheet]]
    return temp


def color_row(row: Tuple[Cell,...], font) -> None:
    """
    sets font of the row to font
    """
    for cell in row:
        cell.font = font


def highlight_missing_lines(wb: Workbook, data: Dict[str, List[Tuple[Cell,...]]]) -> None:
    """
    Hihglights lines in workbook missing from the dataset
    """
    for sheet in wb.worksheets:
        if sheet.title in data.keys():
            for row in sheet.iter_rows(min_row=START_INDEX, max_row=sheet.max_row, max_col=sheet.max_column, min_col=2):
                row_in_data = [cell.value for cell in row] in data[sheet.title]
                if row_in_data:
                    color_row(row, Font(color="2aed18"))  # Green
                else:
                    color_row(row, Font(color="d42815"))  # Red


def get_all_rows(wb: Workbook) -> Dict[str, List[Tuple[Cell,...]]]: #actually not Cell but type(Cell.value)
    """
    Rreturns all rows from workbook into a Dict[sheetname, List[row]]
    """
    dataset: Dict[str, List[Tuple[Cell,...]]] = dict()
    for sheet in wb.worksheets:
        print("maxes", sheet.max_column, sheet.max_row)
        for row in sheet.iter_rows(min_row=START_INDEX, max_row=sheet.max_row, max_col=sheet.max_column, min_col=2):
            if sheet.title not in dataset.keys():
                dataset[sheet.title] = [(cell.value for cell in row)]
            dataset[sheet.title].append((cell.value for cell in row))
    return dataset


def load_excel(path):
    """
    Loads excel table according to path
    """
    if Path(path).exists():
        return load_workbook(filename=path, data_only = False)
    raise Exception("File does not exists")


def get_LVs(path):
    """
    returns names of all property lists
    """
    name_set = set()
    for dir_lv in os.scandir(path):
        name_set.add(dir_lv.name)
    return name_set

if __name__ == "__main__":
    #program()
    s1 = get_LVs("AAAAAAAAAAAAAAAAAA")
    s2 = get_LVs("BBBBBBBBBBBBBBBBBB")

    print([x for x in s1 if x not in s2], [x for x in s2 if x not in s1])
